from flask import Blueprint, flash, redirect, render_template, request, jsonify, make_response, url_for
from flask_login import login_required, current_user
from datetime import datetime, time, timezone
import csv
import io
from sqlalchemy.orm import joinedload
from datetime import timedelta
from collections import Counter

from app import db
from app import csrf
from app.security import roles_required
from app.services.face_service import (
    FaceRecognitionUnavailable,
    clear_per_session_cache,
    preload_known_encodings_for_session,
    recognize_face_from_image_bytes,
    recognize_face_from_frames,
)
from app.services.qr_service import generate_qr_session_token, qr_code_data_uri, parse_student_qr_payload
from app.services.barcode_service import decode_barcode_image_bytes, decode_barcode_data_url
from app.models import Student, User, Attendance, AttendanceSession, Professor, ClassRoom, Subject, LeaveRequest

professor_bp = Blueprint("professor", __name__)


def _escape_pdf_text(value: object) -> str:
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_simple_pdf_bytes(title: str, lines: list[str]) -> bytes:
    content_lines = []
    y = 760
    for line in [title, ""] + lines:
        content_lines.append(f"BT /F1 10 Tf 72 {y} Td ({_escape_pdf_text(line)}) Tj ET")
        y -= 14

    content = "\n".join(content_lines).encode("latin-1", "replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>",
        f"<< /Length {len(content)} >>\nstream\n".encode("latin-1") + content + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("latin-1"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")

    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))

    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("latin-1")
    )
    return bytes(pdf)


def _now_utc_naive():
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _attendance_status_for_session(session, marked_at):
    late_after_minutes = 10
    start = session.starts_at
    if start and start.tzinfo is not None:
        start = start.replace(tzinfo=None)
    if start and marked_at > start + timedelta(minutes=late_after_minutes):
        return "late"
    return "present"


def _active_owned_session(session_id):
    try:
        sess = db.session.get(AttendanceSession, int(session_id))
    except (TypeError, ValueError):
        return None, "session_id invalid", 400

    if not sess:
        return None, "session not found", 404

    if current_user.role_name != "admin":
        prof = current_user.professor_profile
        if not prof or sess.professor_id != prof.id:
            return None, "session not found or unauthorized", 404

    now = _now_utc_naive()
    starts_at = sess.starts_at.replace(tzinfo=None) if sess.starts_at and sess.starts_at.tzinfo else sess.starts_at
    ends_at = sess.ends_at.replace(tzinfo=None) if sess.ends_at and sess.ends_at.tzinfo else sess.ends_at
    if starts_at and starts_at > now:
        return None, "session has not started yet", 409
    if ends_at and ends_at < now:
        return None, "session is already stopped", 409

    return sess, None, None


def _parse_analytics_filters():
    class_id = request.args.get('class_id')
    subject_id = request.args.get('subject_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    start_dt = None
    end_dt = None
    if start_date:
        try:
            start_dt = datetime.combine(datetime.fromisoformat(start_date).date(), time.min)
        except ValueError:
            start_date = ''
    else:
        start_dt = _now_utc_naive() - timedelta(days=30)

    if end_date:
        try:
            end_dt = datetime.combine(datetime.fromisoformat(end_date).date(), time.max)
        except ValueError:
            end_date = ''
    else:
        end_dt = _now_utc_naive()

    return class_id, subject_id, start_date, end_date, start_dt, end_dt


def _analytics_context():
    prof = current_user.professor_profile
    if not prof:
        if current_user.role_name == "admin":
            session_query = AttendanceSession.query
        else:
            return None
    else:
        session_query = AttendanceSession.query.filter_by(professor_id=prof.id)

    class_id, subject_id, start_date, end_date, start_dt, end_dt = _parse_analytics_filters()
    if class_id:
        try:
            session_query = session_query.filter_by(class_id=int(class_id))
        except ValueError:
            class_id = ''
    if subject_id:
        try:
            session_query = session_query.filter_by(subject_id=int(subject_id))
        except ValueError:
            subject_id = ''

    sessions = session_query.order_by(AttendanceSession.starts_at.asc()).all()
    session_ids = [session.id for session in sessions]
    attendance_query = Attendance.query.filter(Attendance.session_id.in_(session_ids)) if session_ids else Attendance.query.filter(False)
    if start_dt:
        attendance_query = attendance_query.filter(Attendance.marked_at >= start_dt)
    if end_dt:
        attendance_query = attendance_query.filter(Attendance.marked_at <= end_dt)
    attendances = attendance_query.order_by(Attendance.marked_at.asc()).all()

    date_counts = Counter(a.marked_at.date().isoformat() for a in attendances if a.marked_at)
    status_counts = Counter(a.attendance_status for a in attendances)
    method_counts = Counter(session.method for session in sessions)
    student_ids = {a.student_id for a in attendances}
    class_course_ids = {session.classroom.course_id for session in sessions if session.classroom}
    enrolled_students = Student.query.filter(Student.course_id.in_(class_course_ids)).count() if class_course_ids else 0
    expected_marks = len(sessions) * enrolled_students if enrolled_students else 0
    total_marks = len(attendances)
    present_marks = status_counts.get('present', 0) + status_counts.get('late', 0)
    attendance_rate = round((present_marks / expected_marks) * 100, 1) if expected_marks else 0
    late_rate = round((status_counts.get('late', 0) / total_marks) * 100, 1) if total_marks else 0

    if not sessions:
        insight = "Selected filters ma koi attendance session nathi."
    elif not attendances:
        insight = "Sessions chhe, pan selected period ma attendance marks nathi."
    elif attendance_rate >= 85:
        insight = "Attendance healthy chhe. Current filters ma participation strong chhe."
    elif attendance_rate >= 60:
        insight = "Attendance moderate chhe. Low attendance students/class review karo."
    else:
        insight = "Attendance low chhe. Follow-up action recommend thay chhe."

    return {
        'filters': {
            'class_id': class_id,
            'subject_id': subject_id,
            'start_date': start_date,
            'end_date': end_date,
        },
        'sessions': sessions,
        'attendances': attendances,
        'date_counts': date_counts,
        'status_counts': status_counts,
        'method_counts': method_counts,
        'stats': {
            'sessions': len(sessions),
            'attendance_marks': total_marks,
            'students_marked': len(student_ids),
            'enrolled_students': enrolled_students,
            'attendance_rate': attendance_rate,
            'late_rate': late_rate,
            'estimated_absences': max(expected_marks - total_marks, 0),
        },
        'insight': insight,
    }


def _fallback_analytics_png(context):
    try:
        from PIL import Image, ImageDraw
    except Exception:
        return None

    width, height = 1000, 380
    image = Image.new('RGB', (width, height), '#181818')
    draw = ImageDraw.Draw(image)
    draw.text((32, 24), 'Attendance Analytics', fill='#FFFFFF')
    draw.text((32, 50), context['insight'], fill='#D9D9E0')

    date_counts = context['date_counts']
    chart_left, chart_top, chart_width, chart_height = 60, 110, 580, 210
    draw.rectangle((chart_left, chart_top, chart_left + chart_width, chart_top + chart_height), outline='#FF7A00')
    if date_counts:
        dates = sorted(date_counts.keys())
        max_count = max(date_counts.values()) or 1
        bar_gap = 8
        bar_width = max(16, int((chart_width - (len(dates) + 1) * bar_gap) / len(dates)))
        for index, date_label in enumerate(dates):
            count = date_counts[date_label]
            bar_height = int((count / max_count) * (chart_height - 40))
            x1 = chart_left + bar_gap + index * (bar_width + bar_gap)
            y1 = chart_top + chart_height - bar_height - 24
            x2 = x1 + bar_width
            y2 = chart_top + chart_height - 24
            draw.rectangle((x1, y1, x2, y2), fill='#FF7A00')
            draw.text((x1, y2 + 4), date_label[-5:], fill='#A1A1AA')
            draw.text((x1, y1 - 16), str(count), fill='#FFFFFF')
    else:
        draw.text((chart_left + 190, chart_top + 95), 'No attendance data', fill='#A1A1AA')

    status_x = 700
    draw.text((status_x, 110), 'Status split', fill='#FFFFFF')
    y = 145
    colors = {'present': '#FF6600', 'late': '#FFAA00', 'absent': '#FF3300', 'leave': '#FF8800'}
    for label, value in context['status_counts'].items():
        draw.rectangle((status_x, y, status_x + 18, y + 18), fill=colors.get(label, '#FFAA00'))
        draw.text((status_x + 28, y), f'{label}: {value}', fill='#D9D9E0')
        y += 34
    if not context['status_counts']:
        draw.text((status_x, y), 'No status data', fill='#A1A1AA')

    draw.text((status_x, 280), f"Rate: {context['stats']['attendance_rate']}%", fill='#FFFFFF')
    draw.text((status_x, 306), f"Late: {context['stats']['late_rate']}%", fill='#FFFFFF')
    draw.text((status_x, 332), f"Absences: {context['stats']['estimated_absences']}", fill='#FFFFFF')

    buffer = io.BytesIO()
    image.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer.getvalue()


@professor_bp.route("/dashboard")
@login_required
@roles_required("professor", "admin")
def dashboard():
    pending_leaves = LeaveRequest.query.filter_by(decision="pending").count()
    return render_template("professor/dashboard.html", pending_leaves=pending_leaves)


@professor_bp.route("/leaves")
@login_required
@roles_required("professor", "admin")
def leave_requests():
    records = (
        LeaveRequest.query.join(Student)
        .join(User)
        .order_by(LeaveRequest.created_at.desc())
        .all()
    )
    return render_template("professor/leaves.html", leaves=records)


@professor_bp.route("/leaves/<int:leave_id>/<decision>", methods=["POST"])
@login_required
@roles_required("professor", "admin")
def decide_leave(leave_id, decision):
    if decision not in {"approved", "rejected"}:
        flash("Invalid leave decision.", "danger")
        return redirect(url_for("professor.leave_requests"))

    leave = db.session.get(LeaveRequest, leave_id)
    if not leave:
        flash("Leave request not found.", "danger")
        return redirect(url_for("professor.leave_requests"))

    leave.decision = decision
    if decision == "approved" and leave.student and leave.starts_on and leave.ends_on:
        start_dt = datetime.combine(leave.starts_on, time.min)
        end_dt = datetime.combine(leave.ends_on, time.max)
        sessions = AttendanceSession.query.filter(
            AttendanceSession.starts_at >= start_dt,
            AttendanceSession.starts_at <= end_dt
        ).all()
        for sess in sessions:
            att = Attendance.query.filter_by(session_id=sess.id, student_id=leave.student.id).first()
            if att:
                att.attendance_status = "leave"
            else:
                db.session.add(Attendance(
                    student_id=leave.student.id,
                    session_id=sess.id,
                    marked_at=start_dt,
                    attendance_status="leave"
                ))

    db.session.commit()
    flash(f"Leave request {decision}.", "success")
    return redirect(url_for("professor.leave_requests"))


@professor_bp.route('/attendance/face', methods=['GET', 'POST'])
@login_required
@roles_required('professor', 'admin')
@csrf.exempt
def face_attendance():
    if request.method == 'GET':
        classes = ClassRoom.query.order_by(ClassRoom.name.asc()).all()
        subjects = Subject.query.order_by(Subject.name.asc()).all()
        return render_template('professor/face_attendance.html', classes=classes, subjects=subjects)

    import base64 as _b64

    # Flexible data extraction: JSON, Form, Raw Data
    data = {}
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        data = request.form.to_dict() if request.form else {}
        if not data and request.data:
            try:
                import json
                data = json.loads(request.data.decode('utf-8'))
            except Exception:
                data = {}

    session_id = data.get('session_id') or request.args.get('session_id')
    if not session_id:
        prof = current_user.professor_profile
        if prof:
            active_sess = AttendanceSession.query.filter_by(
                professor_id=prof.id, method='face', ends_at=None
            ).order_by(AttendanceSession.starts_at.desc()).first()
            if active_sess:
                session_id = active_sess.id

    if not session_id:
        return jsonify({'error': 'Start a session before marking attendance.'}), 400

    sess, session_error, status_code = _active_owned_session(session_id)
    if session_error:
        return jsonify({'error': session_error}), status_code

    frames_raw = data.get('frames')
    img_data   = data.get('image') or data.get('img') or data.get('imageData') or data.get('file')

    # Also check file uploads in request.files
    file_obj = request.files.get('image') or request.files.get('file') or request.files.get('frame')

    if isinstance(frames_raw, str):
        try:
            import json
            frames_raw = json.loads(frames_raw)
        except Exception:
            frames_raw = [frames_raw]

    # ── Decode frames / image ────────────────────────────────────────────────
    frames_bytes = None
    image_bytes  = None

    if file_obj:
        image_bytes = file_obj.read()
    elif frames_raw and isinstance(frames_raw, list) and len(frames_raw) >= 1:
        frames_bytes = []
        for f in frames_raw[:3]:
            if not isinstance(f, str) or not f.strip():
                continue
            b64 = f.split(',', 1)[1] if ',' in f else f
            try:
                frames_bytes.append(_b64.b64decode(b64))
            except Exception:
                pass
        if len(frames_bytes) == 1:
            frames_bytes.append(frames_bytes[0])
        if not frames_bytes:
            return jsonify({'error': 'Invalid frame data.'}), 400
    elif img_data and isinstance(img_data, str) and img_data.strip():
        b64 = img_data.split(',', 1)[1] if ',' in img_data else img_data
        try:
            image_bytes = _b64.b64decode(b64)
        except Exception:
            return jsonify({'error': 'Invalid image data.'}), 400
    else:
        return jsonify({'error': 'No image provided.'}), 400

    # ── Load known encodings ─────────────────────────────────────────────────
    try:
        known = preload_known_encodings_for_session(sess.id)
    except Exception:
        known = None

    # ── Run recognition ──────────────────────────────────────────────────────
    try:
        if frames_bytes:
            result = recognize_face_from_frames(frames_bytes, known=known)
        else:
            result = recognize_face_from_image_bytes(image_bytes, known=known)
    except FaceRecognitionUnavailable as exc:
        return jsonify({'error': str(exc)}), 503
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    if not result:
        return jsonify({'recognized': False})

    # ── Mark attendance ──────────────────────────────────────────────────────
    student_id = result.get('student_id')
    confidence = result.get('confidence')
    student = user = None
    if student_id:
        student = db.session.get(Student, student_id)
        if student:
            user = db.session.get(User, student.user_id)

    persisted = duplicate = False
    attendance_status = None
    if student:
        existing = Attendance.query.filter_by(session_id=sess.id, student_id=student.id).first()
        if not existing:
            now = _now_utc_naive()
            attendance_status = _attendance_status_for_session(sess, now)
            attendance = Attendance(
                student_id=student.id,
                session_id=sess.id,
                marked_at=now,
                attendance_status=attendance_status,
                confidence_score=confidence,
            )
            db.session.add(attendance)
            db.session.commit()
            persisted = True
        else:
            duplicate = True
            attendance_status = existing.attendance_status

    return jsonify({
        'recognized':        True,
        'student_id':        student_id,
        'confidence':        confidence,
        'student_name':      f"{user.first_name} {user.last_name}" if user else None,
        'persisted':         persisted,
        'duplicate':         duplicate,
        'attendance_status': attendance_status,
        'session_id':        sess.id,
    })


@professor_bp.route('/attendance/qr/scan', methods=['POST'])
@login_required
@roles_required('professor', 'admin')
@csrf.exempt
def scan_student_id_qr():
    data = {}
    if request.is_json:
        data = request.get_json(silent=True) or {}
    else:
        data = request.form.to_dict() if request.form else {}

    session_id = data.get('session_id') or request.args.get('session_id')
    qr_data = data.get('qr_data') or data.get('barcode') or data.get('code')
    img_data = data.get('image') or data.get('img') or data.get('imageData')
    file_obj = request.files.get('image') or request.files.get('file') or request.files.get('barcode_file')

    # Automatic Barcode/QR decoding from frame using OpenCV + zxing-cpp if no text provided
    barcode_format = "Text/JS"
    if not qr_data:
        if file_obj:
            b_res = decode_barcode_image_bytes(file_obj.read())
            if b_res:
                qr_data = b_res.text
                barcode_format = b_res.format
        elif img_data and isinstance(img_data, str):
            b_res = decode_barcode_data_url(img_data)
            if b_res:
                qr_data = b_res.text
                barcode_format = b_res.format

    # Auto-detect professor's open QR session if session_id is omitted
    if not session_id:
        prof = current_user.professor_profile
        if prof:
            active_sess = AttendanceSession.query.filter_by(
                professor_id=prof.id, method='qr', ends_at=None
            ).order_by(AttendanceSession.starts_at.desc()).first()
            if active_sess:
                session_id = active_sess.id

    if not session_id:
        return jsonify({'error': 'Start an attendance session first.'}), 400
    if not qr_data:
        return jsonify({'error': 'No QR code or Barcode payload detected.'}), 400

    sess, session_error, status_code = _active_owned_session(session_id)
    if session_error:
        return jsonify({'error': session_error}), status_code

    student_identifier = parse_student_qr_payload(qr_data)
    if not student_identifier:
        student_identifier = str(qr_data).strip()

    student_identifier = str(student_identifier).strip()
    if not student_identifier:
        return jsonify({'error': 'Invalid Student ID / Barcode.'}), 400

    # 0. Match real college ID card barcode (barcode_id column) — HIGHEST PRIORITY
    student = Student.query.filter(
        db.func.lower(Student.barcode_id) == student_identifier.lower()
    ).first()

    # 1. Exact match on roll_number
    if not student:
        student = Student.query.filter(db.func.lower(Student.roll_number) == student_identifier.lower()).first()

    # 2. Partial match on roll_number
    if not student:
        student = Student.query.filter(Student.roll_number.ilike(f"%{student_identifier}%")).first()

    # 3. Match numeric Student ID
    if not student and student_identifier.isdigit():
        student = db.session.get(Student, int(student_identifier))

    # 4. Match student name or email
    if not student:
        student = Student.query.join(User).filter(
            (User.first_name.ilike(f"%{student_identifier}%")) |
            (User.last_name.ilike(f"%{student_identifier}%")) |
            (User.email.ilike(f"%{student_identifier}%"))
        ).first()

    if not student:
        return jsonify({'error': f'Student not found for barcode ({student_identifier}).'}), 404

    user = student.user
    student_name = user.full_name if user else f"Student #{student.id}"

    existing = Attendance.query.filter_by(session_id=sess.id, student_id=student.id).first()
    if existing:
        return jsonify({
            'duplicate': True,
            'student_id': student.id,
            'student_name': student_name,
            'roll_number': student.roll_number,
            'barcode_format': barcode_format,
            'message': 'Attendance already marked.'
        }), 200

    now = _now_utc_naive()
    status = _attendance_status_for_session(sess, now)
    attendance = Attendance(
        student_id=student.id,
        session_id=sess.id,
        marked_at=now,
        attendance_status=status,
        confidence_score=1.0
    )
    db.session.add(attendance)
    db.session.commit()

    total_marked = Attendance.query.filter_by(session_id=sess.id).count()

    return jsonify({
        'success': True,
        'student_id': student.id,
        'student_name': student_name,
        'roll_number': student.roll_number,
        'marked_at': now.strftime('%I:%M:%S %p'),
        'attendance_status': status,
        'barcode_format': barcode_format,
        'total_marked': total_marked
    }), 200


@professor_bp.route('/attendance/qr', methods=['GET'])
@login_required
@roles_required('professor', 'admin')
def qr_attendance():
    classes = ClassRoom.query.order_by(ClassRoom.name.asc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    return render_template('professor/qr_attendance.html', classes=classes, subjects=subjects)


@professor_bp.route('/session/start', methods=['POST'])
@login_required
@roles_required('professor', 'admin')
@csrf.exempt
def start_session():
    data = request.get_json() or {}
    class_id = data.get('class_id')
    subject_id = data.get('subject_id')
    method = data.get('method', 'face')

    prof = current_user.professor_profile
    prof_id = prof.id if prof else (Professor.query.first().id if Professor.query.first() else 1)

    if not class_id:
        return jsonify({'error': 'class_id required'}), 400
    if not subject_id:
        return jsonify({'error': 'subject_id required'}), 400

    try:
        class_obj = db.session.get(ClassRoom, int(class_id))
    except Exception:
        class_obj = None
    if not class_obj:
        return jsonify({'error': 'class_id invalid'}), 400

    try:
        subject_obj = db.session.get(Subject, int(subject_id))
    except Exception:
        subject_obj = None
    if not subject_obj:
        return jsonify({'error': 'subject_id invalid'}), 400

    sess = AttendanceSession(
        method=method,
        starts_at=_now_utc_naive(),
        subject_id=subject_obj.id,
        professor_id=prof_id,
        class_id=class_obj.id,
    )
    db.session.add(sess)
    db.session.commit()
    if method == 'face':
        # preload encodings into memory to speed up realtime recognition for this session
        try:
            preload_known_encodings_for_session(sess.id, force=True)
        except Exception:
            # non-fatal; recognition will fallback to disk
            pass

    response = {'session_id': sess.id}
    if method == 'qr':
        token = generate_qr_session_token(sess.id)
        qr_url = url_for('student.mark_qr_attendance', token=token, _external=True)
        response.update({'qr_url': qr_url, 'qr_image': qr_code_data_uri(qr_url)})
    return jsonify(response)



@professor_bp.route('/session/<int:session_id>/preload', methods=['POST'])
@login_required
@roles_required('professor')
@csrf.exempt
def preload_session_encodings(session_id):
    prof = current_user.professor_profile
    if not prof:
        return jsonify({'error': 'Professor profile missing.'}), 400
    sess = db.session.get(AttendanceSession, session_id)
    if not sess or sess.professor_id != prof.id:
        return jsonify({'error': 'session not found or unauthorized'}), 404
    try:
        preload_known_encodings_for_session(session_id, force=True)
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500
    return jsonify({'preloaded': True})


@professor_bp.route('/session/stop', methods=['POST'])
@login_required
@roles_required('professor', 'admin')
@csrf.exempt
def stop_session():
    data = request.get_json() or {}
    session_id = data.get('session_id')
    if not session_id:
        return jsonify({'error': 'session_id required'}), 400
    try:
        sess = db.session.get(AttendanceSession, int(session_id))
    except Exception:
        sess = None
    if not sess:
        return jsonify({'error': 'session not found'}), 404

    if current_user.role_name != 'admin':
        prof = current_user.professor_profile
        if not prof or sess.professor_id != prof.id:
            return jsonify({'error': 'session not found or unauthorized'}), 404

    sess.ends_at = _now_utc_naive()
    db.session.commit()
    clear_per_session_cache(sess.id)
    return jsonify({'stopped': True})


@professor_bp.route('/sessions')
@login_required
@roles_required('professor', 'admin')
def list_sessions():
    prof = current_user.professor_profile
    if not prof:
        if current_user.role_name == 'admin':
            sessions = AttendanceSession.query.order_by(AttendanceSession.starts_at.desc()).all()
            return render_template('professor/sessions.html', sessions=sessions)
        return jsonify({'error': 'Professor profile missing.'}), 400
    sessions = AttendanceSession.query.filter_by(professor_id=prof.id).order_by(AttendanceSession.starts_at.desc()).all()
    return render_template('professor/sessions.html', sessions=sessions)


@professor_bp.route('/session/<int:session_id>/export')
@login_required
@roles_required('professor', 'admin')
def export_session_csv(session_id):
    prof = current_user.professor_profile
    sess = db.session.get(AttendanceSession, session_id)
    if not sess or (current_user.role_name != 'admin' and (not prof or sess.professor_id != prof.id)):
        return jsonify({'error': 'session not found or unauthorized'}), 404

    attendances = Attendance.query.filter_by(session_id=sess.id).order_by(Attendance.marked_at.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['student_id', 'student_name', 'marked_at', 'attendance_status', 'confidence_score'])
    for a in attendances:
        student = db.session.get(Student, a.student_id)
        name = None
        if student:
            user = db.session.get(User, student.user_id)
            if user:
                name = f"{user.first_name} {user.last_name}"
        writer.writerow([a.student_id, name or '', a.marked_at.isoformat() if a.marked_at else '', a.attendance_status, a.confidence_score])

    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=attendance_session_{sess.id}.csv'
    return resp


@professor_bp.route('/session/<int:session_id>/export.xlsx')
@login_required
@roles_required('professor', 'admin')
def export_session_excel(session_id):
    prof = current_user.professor_profile
    sess = db.session.get(AttendanceSession, session_id)
    if not sess or (current_user.role_name != 'admin' and (not prof or sess.professor_id != prof.id)):
        return jsonify({'error': 'session not found or unauthorized'}), 404

    attendances = Attendance.query.filter_by(session_id=sess.id).order_by(Attendance.marked_at.asc()).all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return jsonify({'error': 'Excel export not available (openpyxl missing).'}), 501

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Session {sess.id}"

    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    headers = ['Student ID', 'Student Name', 'Roll Number', 'Marked At', 'Status', 'Confidence Score']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in attendances:
        student = db.session.get(Student, a.student_id)
        name = ""
        roll = ""
        if student:
            roll = student.roll_number or ""
            user = db.session.get(User, student.user_id)
            if user:
                name = f"{user.first_name} {user.last_name}"
        score = f"{(a.confidence_score * 100):.1f}%" if a.confidence_score is not None else "N/A"
        ws.append([a.student_id, name, roll, a.marked_at.isoformat() if a.marked_at else '', a.attendance_status.upper(), score])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    resp = make_response(buffer.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=attendance_session_{sess.id}.xlsx'
    return resp


@professor_bp.route('/session/<int:session_id>')
@login_required
@roles_required('professor', 'admin')
def session_detail(session_id):
    prof = current_user.professor_profile
    sess = AttendanceSession.query.options(joinedload('professor')).get(session_id)
    if not sess or (current_user.role_name != 'admin' and (not prof or sess.professor_id != prof.id)):
        return jsonify({'error': 'session not found or unauthorized'}), 404

    attendances = Attendance.query.filter_by(session_id=sess.id).order_by(Attendance.marked_at.asc()).all()
    return render_template('professor/session_detail.html', session=sess, attendances=attendances)


@professor_bp.route('/session/<int:session_id>/export.pdf')
@login_required
@roles_required('professor', 'admin')
def export_session_pdf(session_id):
    prof = current_user.professor_profile
    sess = db.session.get(AttendanceSession, session_id)
    if not sess or (current_user.role_name != 'admin' and (not prof or sess.professor_id != prof.id)):
        return jsonify({'error': 'session not found or unauthorized'}), 404

    attendances = Attendance.query.filter_by(session_id=sess.id).order_by(Attendance.marked_at.asc()).all()

    lines = [
        f'Attendance Session {sess.id} - {sess.method}',
        f'Starts: {sess.starts_at.isoformat() if sess.starts_at else ""}',
        f'Ends: {sess.ends_at.isoformat() if sess.ends_at else ""}',
        'Student | Marked At | Confidence',
    ]
    for a in attendances:
        student = db.session.get(Student, a.student_id)
        name = ''
        if student:
            user = db.session.get(User, student.user_id)
            if user:
                name = f"{user.first_name} {user.last_name}"
        lines.append(
            f'{name or "Unknown"} | {a.marked_at.isoformat() if a.marked_at else ""} | {f"{a.confidence_score * 100:.1f}%" if a.confidence_score is not None else "N/A"}'
        )

    pdf_bytes = _build_simple_pdf_bytes('Attendance Session Report', lines)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'attachment; filename=attendance_session_{sess.id}.pdf'
    return resp


@professor_bp.route('/analytics')
@login_required
@roles_required('professor', 'admin')
def analytics_index():
    context = _analytics_context()
    if context is None:
        return jsonify({'error': 'Professor profile missing.'}), 400

    classes = ClassRoom.query.order_by(ClassRoom.name.asc()).all()
    subjects = Subject.query.order_by(Subject.name.asc()).all()
    return render_template(
        'professor/analytics.html',
        classes=classes,
        subjects=subjects,
        stats=context['stats'],
        insight=context['insight'],
        recent_attendances=list(reversed(context['attendances'][-10:])),
        class_id=context['filters']['class_id'],
        subject_id=context['filters']['subject_id'],
        start_date=context['filters']['start_date'],
        end_date=context['filters']['end_date'],
    )


@professor_bp.route('/analytics/export.pdf')
@login_required
@roles_required('professor', 'admin')
def analytics_export_pdf():
    context = _analytics_context()
    if context is None:
        return jsonify({'error': 'Professor profile missing.'}), 400
    lines = [
        'Automated Attendance Analytics Report',
        f"Sessions: {context['stats']['sessions']} | Total Marks: {context['stats']['attendance_marks']} | Rate: {context['stats']['attendance_rate']}%",
        f"Insight: {context['insight'][:110]}",
        f"Estimated Absences: {context['stats']['estimated_absences']} | Late Rate: {context['stats']['late_rate']}% | Enrolled: {context['stats']['enrolled_students']}",
        'Recent Attendance Records',
    ]
    for att in reversed(context['attendances'][-20:]):
        student_name = att.student.user.full_name if att.student and att.student.user else f"Student #{att.student_id}"
        lines.append(f"{student_name[:35]} | {att.attendance_status.upper()} | {att.marked_at.strftime('%Y-%m-%d %H:%M') if att.marked_at else ''}")

    pdf_bytes = _build_simple_pdf_bytes('Attendance Analytics Report', lines)
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = 'attachment; filename=attendance_analytics.pdf'
    return resp
